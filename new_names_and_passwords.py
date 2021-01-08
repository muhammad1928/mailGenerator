from openpyxl import load_workbook
import random
from functions import Password

birth_months = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
# alphabet_uppercase = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'X', 'Y', 'Z']
# alphabet_lowercase = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u', 'v', 'x', 'y', 'z']
# numbers = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]

# data file we feed
filename = "1000 names.xlsx"
# data file we will create
dest_filename = "test3.xlsx"

# getting names as a list from our data
list_names = []
list_surnames = []

# loading the workbook
wb = load_workbook(filename)
# activate the sheet
ws = wb.active

# getting the names and surnames
column_name = ws['A']
column_surname = ws['B']

# going through all the names and adding them to the list
for cell in column_name:
    a = cell.value
    list_names.append(a)

# going through all the surnames and adding them to the list
for cell in column_surname:
    b = cell.value
    list_surnames.append(b)


# getting the lenght of our list
lenght_names = len(list_names)
lenght_surname = len(list_surnames)

# print(len(list_names))
# print(len(list_surnames))
# print(list_names)
# print(list_surnames)

# creating a new sheet called NewData
ws2 = wb.create_sheet("NewData")  # insert at the end (default)

# getting our lenght of the list
max_row = len(list_names)

# writing the headers
ws2['A1'] = "name"
ws2['B1'] = "surname"
ws2['C1'] = "Birthday day"
ws2['D1'] = "Birthday month"
ws2['E1'] = "Birthday year"
ws2['F1'] = "Mail"
ws2['G1'] = "password"


# we are calling the names one by one randomly
# starting from one because number 2 are the headers
for row in range(2, max_row):

    # giving x and y random names and surnames from excel
    x = random.randint(2, max_row-1)
    y = random.randint(2, max_row-1)
    # getting names randomly
    name = list_names[x]
    surname = list_surnames[y]
    # picking random month
    birth_month = random.choice(birth_months)

    # if the month is not february, get a number from 1 to 30
    if birth_month != birth_months[1]:
        birth_day = random.randint(1, 30)

    # if the month is february, get a number from 1 to 28
    else:
        birth_day = random.randint(1, 28)

    # choose a random year between 1970 - 2003
    year = random.randint(1970, 2003)
    # create a random mail adress
    mail_numbers = random.randint(1000, 9999)
    mail_new = name + "." + surname + str(mail_numbers)
    # print(mail_new)
    password_new = Password.generate()
    # print(password_new)
    # writing the values
    column_a = ws2.cell(row=row, column=1)
    column_a.value = name
    column_a = ws2.cell(row=row, column=2)
    column_a.value = surname
    column_a = ws2.cell(row=row, column=3)
    column_a.value = birth_day
    column_a = ws2.cell(row=row, column=4)
    column_a.value = birth_month
    column_a = ws2.cell(row=row, column=5)
    column_a.value = year
    column_a = ws2.cell(row=row, column=6)
    column_a.value = mail_new
    column_a = ws2.cell(row=row, column=7)
    column_a.value = password_new

# saving the workbook
wb.save(dest_filename)
